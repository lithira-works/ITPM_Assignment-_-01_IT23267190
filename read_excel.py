import openpyxl
import json

excel_file = 'Assignment 1 - Test cases.xlsx'
wb = openpyxl.load_workbook(excel_file)

print('Available sheets:', wb.sheetnames)
print()

# Get the active sheet
ws = wb.active
print(f'Active sheet: {ws.title}')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')
print()

# Print first 30 rows to understand the structure
print('First 30 rows of data:')
print('='*120)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    print(f'Row {i}: {row}')

print()
print('='*120)
print(f'\nTotal test cases: {ws.max_row - 1}')  # Assuming row 1 is header
